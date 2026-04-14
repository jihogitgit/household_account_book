import msoffcrypto
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base = '/Users/mw/prodect/통장/'
files = [
    ('생활비', '토스뱅크_거래내역 _생활비.xlsx'),
    ('경조사', '토스뱅크_거래내역_경조사.xlsx'),
    ('급여통장', '토스뱅크_거래내역_급여통장.xlsx'),
    ('비상금', '토스뱅크_거래내역_비상금.xlsx'),
]
password = '911017'

# ── 색상 (6자리 RGB) ──────────────────────────────────────────────────
CAT_BG = {
    '수입':    'D6F5D6',
    '고정지출': 'DCE8FF',
    '변동지출': 'FFF3CD',
    '경조사':  'FDE0E0',
    '내부이체': 'F0F0F0',
    '기타':    'FAFAFA',
    '미분류':  'FF9999',
}
CAT_DARK = {
    '수입':    'A8E6A8',
    '고정지출': 'B0C8F0',
    '변동지출': 'FFE08A',
    '경조사':  'F4B8B8',
    '내부이체': 'DDDDDD',
    '기타':    'EEEEEE',
    '미분류':  'FF6666',
}
HDR  = '2D4A6B'
DARK = '1A2E4A'
FIXED_HDR  = '1A3A6E'   # 고정지출 헤더 (진한 파랑)
VAR_HDR    = 'B45309'   # 변동지출 헤더 (진한 주황)
FIXED_SECT = 'C8DCFA'   # 고정지출 섹션 배경
VAR_SECT   = 'FEE6A0'   # 변동지출 섹션 배경
INCOME_HDR = '1D5C2C'   # 수입 헤더

# IsFixed 대상 소분류
FIXED_SUBCATS = {
    '월세/주거', '주거/통신', '보험', '연금', '적금/저축',
    '공과금', '용돈', '정기구독',
}


def pf(hex6):
    return PatternFill('solid', fgColor=hex6)


def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_cell(ws, row, col, val, width=None, bg=HDR, fg='FFFFFF', size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = pf(bg)
    c.font = Font(bold=True, color=fg, size=size)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def money_cell(ws, row, col, val, bg='FFFFFF', bold=False, size=9, fg_override=None):
    v = pd.to_numeric(val, errors='coerce') if not isinstance(val, (int, float)) else val
    v = None if (v is not None and pd.isna(v)) else v
    c = ws.cell(row=row, column=col, value=v if v else '')
    c.fill = pf(bg)
    c.border = border()
    c.number_format = '#,##0'
    c.alignment = Alignment(horizontal='right', vertical='center')
    if fg_override:
        color = fg_override
    elif isinstance(v, float):
        color = 'CC0000' if v < 0 else ('1B7F2D' if v > 0 else '888888')
    else:
        color = '000000'
    c.font = Font(bold=bold, color=color, size=size)
    return c


def txt_cell(ws, row, col, val, bg='FFFFFF', bold=False, color='000000',
             size=9, align='left'):
    c = ws.cell(row=row, column=col, value=str(val) if pd.notna(val) and val != '' else '')
    c.fill = pf(bg)
    c.font = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = border()
    return c


# ── 카테고리 규칙 ─────────────────────────────────────────────────────
# (대분류, 소분류, IsFixed, 조건함수) 순서 중요 — 위에서 먼저 매칭
rules = [
    # ── 수입 ──────────────────────────────────────────────────────────
    ('수입', '급여', False,
        lambda r: r['거래 유형'] == '입금' and r['적요'] in ['유지호', '김채현']),
    ('수입', '이자/캐시백', False,
        lambda r: r['거래 유형'] in ['이자입금', '프로모션입금']),
    ('수입', '모임통장 정산', False,
        lambda r: r['거래 유형'] == '모임원송금'
                  and float(str(r['거래 금액']).replace(',', '')) > 0),

    # ── 내부이체 ──────────────────────────────────────────────────────
    ('내부이체', '생활비 이체', False,
        lambda r: r['적요'] == '생활비'
                  and r['거래 유형'] in ['입금', '자동이체', '내계좌간자동이체', '출금']),
    ('내부이체', '여행비 이체', False,
        lambda r: r['적요'] == '여행비'
                  and r['거래 유형'] in ['입금', '자동이체', '내계좌간자동이체', '출금']),
    ('내부이체', '추가 저축', False,
        lambda r: r['적요'] == '추가 저축'),
    ('내부이체', '계좌간 이체', False,
        lambda r: r['거래 유형'] == '내계좌간자동이체'),

    # ── 고정지출: 주거/통신 (KT/LG/SK/통신 포함 → 최우선 재분류) ────
    ('고정지출', '주거/통신', True,
        lambda r: bool(
            pd.Series([str(r['적요'])]).str.contains(
                r'KT|LG|SK|통신', case=False, na=False
            ).iloc[0]
        )),

    # ── 고정지출 ──────────────────────────────────────────────────────
    ('고정지출', '월세/주거', True,   lambda r: r['적요'] == '월세/이자'),
    ('고정지출', '보험', True,        lambda r: '보험' in str(r['적요'])),
    ('고정지출', '연금', True,        lambda r: '연금' in str(r['적요'])),
    ('고정지출', '적금/저축', True,
        lambda r: any(x in str(r['적요']) for x in ['적금', '청약', '청년도약'])),
    ('고정지출', '교통비', False,     # 교통만 단독 (통신 제외)
        lambda r: r['적요'] in ['채현 교통비'] and r['거래 유형'] == '자동이체'),
    ('고정지출', '용돈', True,        lambda r: '용돈' in str(r['적요'])),
    ('고정지출', '공과금', True,      lambda r: r['거래 유형'] == '지로출금'),

    # ── 경조사 ────────────────────────────────────────────────────────
    ('경조사', '경조사', False,       lambda r: r['_통장'] == '경조사'),

    # ── 변동지출: 정기구독 (IsFixed=True) ────────────────────────────
    ('고정지출', '정기구독', True,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      'JITTER', 'MILS', '넷플릭스', 'netflix', '유튜브', 'youtube',
                      'spotify', '스포티파이', '왓챠', '웨이브',
                  ])),

    # ── 변동지출 ──────────────────────────────────────────────────────
    ('변동지출', '배달', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in ['우아한형제들', '쿠팡이츠'])),
    ('변동지출', '식비', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '구의문복합식당', '겐로쿠우동', '동대문곱창', '마마쿡', '멕시카나', '멘노아지',
                      '바다어묵나라', '삼진스트라이크존', '석문어', '속초오징어', '식물원복합식당',
                      '짬뽕지존', '탄토탄토', '해운대대구탕', '호미스피자', '훼미리손칼국수',
                      '유미분김밥', '란영양', '느굿', '담벼락핫도그', '레드브릭스모크하우스',
                      '레이지아워', '롱메', '몽마르카부덴', '소소달', '소풍', '아리계곡', '야생과',
                      '엠엠씨', '원효로105', '정안', '제이지푸드시스템', '제주돔베고기집',
                      '순천미향', '애월장인', '슬로보트', '심학산도토리',
                      '이마트', '정성마트', '이편한마트', '이편한 정육점', '롯데프레시', '샘터마트',
                      '유니드라멘', '명랑핫도그', '옥이네수산', '미래',
                  ])),
    ('변동지출', '카페/음료', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '커피', '카페', '공차', '더벤티', '매머드', '메가엠지씨', '잔물결', '컴포즈',
                      '씨미트', '가배도', '브루브루', '마노커피', '뚜레쥬르', '오투', '베이커리',
                      '베이글', '런던베이글', '빵', '과자점', '제과', '젤라또', '배스킨', '팔레트',
                      '신세계제과', '우리쌀빵', '윤숲', '온더브레드', '피코야', '투썸플레이스',
                  ])),
    ('변동지출', '편의점', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      'GS25', 'gs25', '지에스25', '씨유', 'CU', '세븐일레븐', '이마트24',
                  ])),
    ('변동지출', '쇼핑', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '쿠팡', '다이소', '아트박스', '에스에스지', '네이버페이', '현대백',
                      '에이케이플라자', '롯데물산', '29cm', '오늘의집', '마켓컬리', '후추포인트',
                  ])),
    ('변동지출', '뷰티/미용', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in ['올리브영', '엘라스틴', '와이즐리'])),
    ('변동지출', '의료/약국', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '약국', '병원', '의원', '린여성', '네이처스파',
                  ])),
    ('변동지출', '문화/여가', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      'CGV', '출판도시', '스파랜드', '스너글리', '시소',
                  ])),
    ('변동지출', '교통', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '고속버스', '코레일', '티머니', '택시',
                  ])),
    ('변동지출', '여행', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in [
                      '놀유니버스', '리조트', '제주', '애월', '서귀포', '주유소', '휴게소',
                  ])),
    ('변동지출', '반려동물', False,
        lambda r: r['거래 유형'] == '체크카드결제'
                  and any(x in str(r['적요']) for x in ['길고양이', '마르못'])),
    ('변동지출', '현금(ATM)', False,   lambda r: r['거래 유형'] == 'ATM출금'),

    # ── 기타 ──────────────────────────────────────────────────────────
    ('기타', '모임 출금', False,
        lambda r: r['거래 유형'] == '모임원송금'
                  and float(str(r['거래 금액']).replace(',', '')) < 0),
    ('기타', '개인 이체', False,
        lambda r: r['거래 유형'] in ['입금', '출금']),
]


def categorize(row):
    for 대분류, 소분류, is_fixed, cond in rules:
        try:
            if cond(row):
                return 대분류, 소분류, is_fixed
        except Exception:
            pass
    return '미분류', '미분류', False


def decrypt(path):
    with open(path, 'rb') as f:
        of = msoffcrypto.OfficeFile(f)
        of.load_key(password=password)
        dec = io.BytesIO(); of.decrypt(dec); dec.seek(0)
    return dec


def load_data():
    all_rows = []
    for label, f in files:
        dec = decrypt(base + f)
        df = pd.read_excel(dec, engine='openpyxl', header=8)
        df['_통장'] = label
        all_rows.append(df)
    df = pd.concat(all_rows, ignore_index=True)
    df = df.dropna(subset=['거래 일시'])
    result = df.apply(lambda r: pd.Series(categorize(r)), axis=1)
    df['대분류'] = result[0]
    df['소분류'] = result[1]
    df['IsFixed'] = result[2]
    df['거래금액'] = pd.to_numeric(df['거래 금액'], errors='coerce')
    df['거래일시'] = pd.to_datetime(df['거래 일시'], errors='coerce')
    df['날짜'] = df['거래일시'].dt.date
    df['연월'] = df['거래일시'].dt.to_period('M').astype(str)
    return df


# ── 시트 1: 전체내역 ──────────────────────────────────────────────────
def sheet_all(wb, df):
    ws = wb.create_sheet('📋 전체내역')
    ws.freeze_panes = 'A2'

    cols = [('날짜', 14), ('적요', 30), ('거래유형', 13), ('통장', 10),
            ('대분류', 11), ('소분류', 13), ('IsFixed', 9),
            ('거래금액', 14), ('잔액', 16), ('메모', 20)]
    src  = ['날짜', '적요', '거래 유형', '_통장', '대분류', '소분류', 'IsFixed',
            '거래금액', '거래 후 잔액', '메모']

    for ci, (h, w) in enumerate(cols, 1):
        hdr_cell(ws, 1, ci, h, width=w)
    ws.row_dimensions[1].height = 22

    for ri, (_, row) in enumerate(df.sort_values('거래일시', ascending=False).iterrows(), 2):
        bg = CAT_BG.get(row['대분류'], 'FAFAFA')
        for ci, col in enumerate(src, 1):
            val = row.get(col, '')
            if col in ('거래금액', '거래 후 잔액'):
                money_cell(ws, ri, ci, val, bg=bg)
            elif col == 'IsFixed':
                fixed_val = bool(row.get('IsFixed', False))
                label = '✔ 고정' if fixed_val else ''
                txt_cell(ws, ri, ci, label, bg=bg,
                         bold=fixed_val, color='1A3A6E' if fixed_val else '999999',
                         align='center')
            else:
                txt_cell(ws, ri, ci, val, bg=bg)
        ws.row_dimensions[ri].height = 16

    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'


# ── 시트 2: 날짜별 ───────────────────────────────────────────────────
def sheet_daily(wb, df):
    ws = wb.create_sheet('📅 날짜별')
    ws.freeze_panes = 'A2'

    ddf = df[~df['대분류'].isin(['내부이체'])].sort_values('거래일시', ascending=True)

    cols = [('날짜', 14), ('적요', 30), ('대분류', 11), ('소분류', 13),
            ('통장', 10), ('IsFixed', 9), ('수입', 14), ('지출', 14), ('잔액', 16)]
    for ci, (h, w) in enumerate(cols, 1):
        hdr_cell(ws, 1, ci, h, width=w)
    ws.row_dimensions[1].height = 22

    prev_date = None
    ri = 2
    for _, row in ddf.iterrows():
        cur_date = str(row['날짜'])
        is_new = cur_date != prev_date
        bg = CAT_BG.get(row['대분류'], 'FAFAFA')
        date_bg = 'D6E4F7' if is_new else bg

        amt = row['거래금액']
        income  = amt if isinstance(amt, float) and amt > 0 else ''
        expense = amt if isinstance(amt, float) and amt < 0 else ''
        fixed_val = bool(row.get('IsFixed', False))

        txt_cell(ws, ri, 1, cur_date if is_new else '', bg=date_bg,
                 bold=is_new, color='2D4A6B' if is_new else '888888', align='center')
        txt_cell(ws, ri, 2, row.get('적요', ''), bg=bg)
        txt_cell(ws, ri, 3, row.get('대분류', ''), bg=bg)
        txt_cell(ws, ri, 4, row.get('소분류', ''), bg=bg)
        txt_cell(ws, ri, 5, row.get('_통장', ''), bg=bg)
        txt_cell(ws, ri, 6, '✔ 고정' if fixed_val else '', bg=bg,
                 bold=fixed_val, color='1A3A6E' if fixed_val else '999999', align='center')
        money_cell(ws, ri, 7, income,  bg=bg)
        money_cell(ws, ri, 8, expense, bg=bg)
        money_cell(ws, ri, 9, row.get('거래 후 잔액', ''), bg=bg)

        ws.row_dimensions[ri].height = 16
        prev_date = cur_date
        ri += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'


# ── 시트 3: 카테고리별 요약 (고정/변동 구분) ─────────────────────────
def sheet_category(wb, df):
    ws = wb.create_sheet('🗂️ 카테고리별')

    cdf = df[~df['대분류'].isin(['내부이체'])].copy()
    months = sorted(cdf['연월'].unique())
    n_cols = 3 + len(months) + 1   # 대분류 | 소분류 | IsFixed | months... | 합계

    # ── 타이틀 ──
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    c = ws.cell(row=1, column=1, value='카테고리별 지출·수입 요약 (고정/변동 구분)')
    c.fill = pf(DARK); c.font = Font(bold=True, color='FFFFFF', size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── 헤더 ──
    hdr_cell(ws, 2, 1, '대분류', width=12)
    hdr_cell(ws, 2, 2, '소분류', width=14)
    hdr_cell(ws, 2, 3, 'IsFixed', width=9)
    for mi, m in enumerate(months, 4):
        hdr_cell(ws, 2, mi, m, width=13)
    hdr_cell(ws, 2, n_cols, '전체 합계', width=15)
    ws.row_dimensions[2].height = 20

    ri = 3
    totals_m = {m: 0.0 for m in months}
    grand = 0.0

    # ── 섹션 순서: 수입 → 고정지출(IsFixed) → 변동지출 → 경조사 → 기타 ──
    SECTIONS = [
        ('수입',    '수입',    INCOME_HDR, 'A8E6A8', False),
        ('고정지출', '고정지출', FIXED_HDR,  FIXED_SECT, True),
        ('변동지출', '변동지출', VAR_HDR,    VAR_SECT, False),
        ('경조사',  '경조사',  'B45309',   'FDE0E0', False),
        ('기타',    '기타',    '555555',   'EEEEEE', False),
        ('미분류',  '미분류',  'CC0000',   'FF9999', False),
    ]

    for cat, label, hdr_color, sect_bg, is_fixed_section in SECTIONS:
        subdf = cdf[cdf['대분류'] == cat]
        if subdf.empty:
            continue

        # 섹션 헤더 행
        ws.merge_cells(f'A{ri}:{get_column_letter(n_cols)}{ri}')
        section_label = label
        if cat == '고정지출':
            section_label = '■ 고정 지출 (IsFixed: True)'
        elif cat == '변동지출':
            section_label = '■ 변동 지출 (IsFixed: False)'
        elif cat == '수입':
            section_label = '■ 수입'
        c = ws.cell(row=ri, column=1, value=section_label)
        c.fill = pf(hdr_color)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = border()
        # 나머지 병합 셀 테두리
        for ci in range(2, n_cols + 1):
            ws.cell(row=ri, column=ci).fill = pf(hdr_color)
            ws.cell(row=ri, column=ci).border = border()
        ws.row_dimensions[ri].height = 22
        ri += 1

        # 대분류 합계 행
        cat_m = {m: subdf[subdf['연월'] == m]['거래금액'].sum() for m in months}
        cat_g = subdf['거래금액'].sum()

        ws.merge_cells(f'A{ri}:B{ri}')
        txt_cell(ws, ri, 1, f'▶ {cat} 합계', bg=CAT_DARK.get(cat, 'EEEEEE'),
                 bold=True, size=10)
        ws.cell(row=ri, column=2).fill = pf(CAT_DARK.get(cat, 'EEEEEE'))
        ws.cell(row=ri, column=2).border = border()
        fixed_label = '✔ 고정' if is_fixed_section else ''
        txt_cell(ws, ri, 3, fixed_label, bg=CAT_DARK.get(cat, 'EEEEEE'),
                 bold=True, color='1A3A6E', align='center')
        for mi, m in enumerate(months, 4):
            money_cell(ws, ri, mi, cat_m[m] if cat_m[m] else '',
                       bg=CAT_DARK.get(cat, 'EEEEEE'), bold=True, size=9)
        money_cell(ws, ri, n_cols, cat_g if cat_g else '',
                   bg=CAT_DARK.get(cat, 'EEEEEE'), bold=True, size=10)
        ws.row_dimensions[ri].height = 20
        ri += 1

        # 소분류별 행
        for sub in sorted(subdf['소분류'].unique()):
            sdf = subdf[subdf['소분류'] == sub]
            sub_is_fixed = bool(sdf['IsFixed'].mode().iloc[0]) if not sdf.empty else False
            bg = CAT_BG.get(cat, 'FAFAFA')

            txt_cell(ws, ri, 1, '', bg=bg)
            txt_cell(ws, ri, 2, sub, bg=bg, size=9)
            fixed_tag = '✔ 고정' if sub_is_fixed else ''
            txt_cell(ws, ri, 3, fixed_tag, bg=bg,
                     bold=sub_is_fixed, color='1A3A6E' if sub_is_fixed else 'BBBBBB',
                     align='center', size=9)

            sub_g = 0.0
            for mi, m in enumerate(months, 4):
                v = sdf[sdf['연월'] == m]['거래금액'].sum()
                sub_g += v
                money_cell(ws, ri, mi, v if v else '', bg=bg, size=9)
            money_cell(ws, ri, n_cols, sub_g if sub_g else '',
                       bg=bg, bold=True, size=9)
            ws.row_dimensions[ri].height = 16
            ri += 1

        # 실질 합산 (수입/고정/변동/경조사만)
        if cat in ('수입', '고정지출', '변동지출', '경조사'):
            for m in months:
                totals_m[m] += cat_m[m]
            grand += cat_g

        ri += 1  # 섹션 사이 빈 행

    # ── 최종 합계 ──
    ws.merge_cells(f'A{ri}:B{ri}')
    txt_cell(ws, ri, 1, '실질 수지 합계', bg=DARK, bold=True,
             color='FFFFFF', size=11, align='center')
    ws.cell(row=ri, column=2).fill = pf(DARK); ws.cell(row=ri, column=2).border = border()
    txt_cell(ws, ri, 3, '', bg=DARK)
    for mi, m in enumerate(months, 4):
        v = totals_m[m]
        money_cell(ws, ri, mi, v if v else '', bg=DARK, bold=True, size=10,
                   fg_override='FFFFFF')
    money_cell(ws, ri, n_cols, grand, bg=DARK, bold=True, size=12,
               fg_override='FFFFFF')
    ws.row_dimensions[ri].height = 26


# ── 시트 4: 월별 요약 (고정/변동 구분) ──────────────────────────────
def sheet_monthly(wb, df):
    ws = wb.create_sheet('📆 월별요약')

    mdf = df[~df['대분류'].isin(['내부이체'])].copy()
    months = sorted(mdf['연월'].unique())

    ws.merge_cells('A1:J1')
    c = ws.cell(row=1, column=1, value='월별 수입 / 지출 요약')
    c.fill = pf(DARK); c.font = Font(bold=True, color='FFFFFF', size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    hdrs = [
        ('연월', 12), ('총수입', 14),
        ('고정지출 계', 14), ('월세/주거', 12), ('주거/통신', 12),
        ('보험', 10), ('연금', 10), ('적금/저축', 12), ('공과금', 10),
        ('용돈', 10), ('정기구독', 11),
        ('변동지출 계', 14), ('식비', 10), ('카페/음료', 11), ('배달', 10),
        ('편의점', 10), ('쇼핑', 10), ('뷰티/미용', 11), ('의료/약국', 11),
        ('문화/여가', 11), ('교통', 10), ('여행', 10), ('반려동물', 11), ('현금(ATM)', 11),
        ('경조사', 10), ('순수지', 14),
    ]
    for ci, (h, w) in enumerate(hdrs, 1):
        # 고정지출 헤더는 파랑, 변동지출 헤더는 주황
        if h in ('고정지출 계', '월세/주거', '주거/통신', '보험', '연금',
                 '적금/저축', '공과금', '용돈', '정기구독'):
            bg = FIXED_HDR
        elif h in ('변동지출 계', '식비', '카페/음료', '배달', '편의점', '쇼핑',
                   '뷰티/미용', '의료/약국', '문화/여가', '교통', '여행', '반려동물', '현금(ATM)'):
            bg = VAR_HDR
        elif h == '순수지':
            bg = DARK
        else:
            bg = HDR
        hdr_cell(ws, 2, ci, h, width=w, bg=bg)
    ws.row_dimensions[2].height = 22

    fixed_subs = ['월세/주거', '주거/통신', '보험', '연금', '적금/저축',
                  '공과금', '용돈', '정기구독', '교통비']
    var_subs   = ['식비', '카페/음료', '배달', '편의점', '쇼핑', '뷰티/미용',
                  '의료/약국', '문화/여가', '교통', '여행', '반려동물', '현금(ATM)']

    for ri, m in enumerate(months, 3):
        mrow = mdf[mdf['연월'] == m]
        bg = 'F0F5FF' if ri % 2 == 0 else 'FFFFFF'

        def sub_sum(sub):
            v = mrow[mrow['소분류'] == sub]['거래금액'].sum()
            return v if v else ''

        fixed_total = mrow[mrow['대분류'] == '고정지출']['거래금액'].sum()
        var_total   = mrow[mrow['대분류'] == '변동지출']['거래금액'].sum()
        income      = mrow[mrow['대분류'] == '수입']['거래금액'].sum()
        event       = mrow[mrow['대분류'] == '경조사']['거래금액'].sum()
        net         = income + fixed_total + var_total + event

        vals = [
            m, income, fixed_total,
            sub_sum('월세/주거'), sub_sum('주거/통신'),
            sub_sum('보험'), sub_sum('연금'), sub_sum('적금/저축'),
            sub_sum('공과금'), sub_sum('용돈'), sub_sum('정기구독'),
            var_total,
            sub_sum('식비'), sub_sum('카페/음료'), sub_sum('배달'),
            sub_sum('편의점'), sub_sum('쇼핑'), sub_sum('뷰티/미용'), sub_sum('의료/약국'),
            sub_sum('문화/여가'), sub_sum('교통'), sub_sum('여행'), sub_sum('반려동물'), sub_sum('현금(ATM)'),
            event, net,
        ]

        for ci, v in enumerate(vals, 1):
            if ci == 1:
                txt_cell(ws, ri, ci, v, bg=bg, bold=True, size=10, align='center')
            else:
                is_total = ci in (3, 12, len(vals))
                money_cell(ws, ri, ci, v, bg=bg, bold=is_total, size=9 if not is_total else 10)
        ws.row_dimensions[ri].height = 18

    # 합계 행
    last_data = 2 + len(months)
    total_ri = last_data + 1
    txt_cell(ws, total_ri, 1, '합 계', bg=DARK, bold=True,
             color='FFFFFF', size=10, align='center')
    ws.cell(row=total_ri, column=1).fill = pf(DARK)
    for ci in range(2, len(hdrs) + 1):
        col_l = get_column_letter(ci)
        c = ws.cell(row=total_ri, column=ci,
                    value=f'=SUM({col_l}3:{col_l}{last_data})')
        c.fill = pf(DARK)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.border = border()
    ws.row_dimensions[total_ri].height = 22

    ws.freeze_panes = 'B3'


# ── 메인 ────────────────────────────────────────────────────────────
def main():
    print('데이터 로딩...')
    df = load_data()
    print(f'총 {len(df)}건 로드 완료')

    # 변경 사항 확인
    telecom = df[df['소분류'] == '주거/통신']
    print(f'\n[주거/통신] 재분류: {len(telecom)}건')
    print(telecom[['적요', '거래 유형', '_통장', '거래금액']].to_string())

    fixed = df[df['IsFixed'] == True]
    print(f'\n[IsFixed=True] 총: {len(fixed)}건')
    print(fixed.groupby('소분류')['거래금액'].sum().to_string())

    wb = Workbook()
    wb.remove(wb.active)

    print('\n시트 생성 중...')
    sheet_all(wb, df)
    sheet_daily(wb, df)
    sheet_category(wb, df)
    sheet_monthly(wb, df)

    out = base + '가계부_리포트.xlsx'
    wb.save(out)
    print(f'\n저장 완료: {out}')
    for ws in wb.worksheets:
        print(f'  - {ws.title}')


if __name__ == '__main__':
    main()
